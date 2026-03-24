# -*- coding: utf-8 -*-
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

# 读取表1（匹配结果）
df1 = pd.read_excel(r'C:\Users\lenovo\Desktop\表1_匹配结果.xlsx')

# 读取表2（机具数量）- 综合查询sheet
xl = pd.ExcelFile(r'C:\Users\lenovo\Desktop\桂林银行南宁分行机具数量（截至20221105）(1).xlsx')
df2 = pd.read_excel(xl, sheet_name=1, header=None)

# 设置列名
df2.columns = ['网点名称', '设备类型', '品牌', '数量', 'col4', 'col5', 'col6', '品牌2', '维护方']

# 填充网点名称（每组第一行有网点名，后续行是NaN）
df2['网点名称'] = df2['网点名称'].ffill()

# 过滤有效数据（去掉标题行）
df2 = df2[df2['设备类型'].notna()]

# 按网点统计设备数量
df2_counts = df2.groupby('网点名称')['数量'].sum().reset_index()
df2_counts.columns = ['网点名称', '机具数量_表2']

# 按网点统计表1的设备数量
df1_counts = df1.groupby('网点名称')['设备名称'].count().reset_index()
df1_counts.columns = ['网点名称', '设备数量_表1']

# 合并对比
result = pd.merge(df1_counts, df2_counts, on='网点名称', how='outer')
result['差异'] = result['设备数量_表1'] - result['机具数量_表2']
result['差异'] = result['差异'].fillna(0)

# 标记有差异的行
result['有差异'] = result['差异'] != 0

print('校对结果统计:')
print(f'总网点数: {len(result)}')
print(f'有差异的网点数: {result["有差异"].sum()}')
print()
print('有差异的网点:')
diff_rows = result[result['有差异']]
print(diff_rows)

# 创建Excel（带红色标记）
wb = Workbook()
ws = wb.active

# 表头
headers = ['网点名称', '表1设备数量', '表2机具数量', '差异', '校对结果', '备注']
for col, header in enumerate(headers, 1):
    ws.cell(row=1, column=col, value=header)

# 写入数据
for row_idx, row in result.iterrows():
    r = row_idx + 2
    ws.cell(row=r, column=1, value=row['网点名称'])
    ws.cell(row=r, column=2, value=row.get('设备数量_表1', 0) if pd.notna(row.get('设备数量_表1')) else 0)
    ws.cell(row=r, column=3, value=row.get('机具数量_表2', 0) if pd.notna(row.get('机具数量_表2')) else 0)
    ws.cell(row=r, column=4, value=int(row['差异']))
    ws.cell(row=r, column=5, value='有差异' if row['有差异'] else '一致')
    ws.cell(row=r, column=6, value='')

    # 如果有差异，标记整行为红色
    if row['有差异']:
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        for col in range(1, 7):
            ws.cell(row=r, column=col).fill = red_fill

# 调整列宽
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 20

wb.save(r'C:\Users\lenovo\Desktop\校对结果.xlsx')
print()
print('校对结果已保存到: C:\\Users\\lenovo\\Desktop\\校对结果.xlsx')
